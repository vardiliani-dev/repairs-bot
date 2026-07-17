import logging
import asyncio
import json
import re
import os
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationHandlerStop,
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)
import gspread
from google.oauth2.service_account import Credentials
import anthropic

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РќРђР›РђРЁРўРЈР’РђРќРќРЇ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
BOT_TOKEN  = os.environ.get("BOT_TOKEN", "")
SHEET_ID   = "1Nq-RKRAF16ZOs2gq7RS7IZ5-6PFsxZrNT4MkJ75dJ9U"
CREDS_FILE = "create-497113-eed86744057e.json"

MANAGER_IDS = [805571381, 692989160, 321443422]
MANAGER_NAMES = {
    805571381: "РћР»РµРєСЃР°РЅРґСЂ",
    692989160: "Р’РёС‚Р°Р»С–Р№",
    321443422: "Рћ.Рћ.",
}
DIRECTOR_ID   = 299617056
ACCOUNTANT_ID = 5030873843

SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]

# РЎРїРёСЃРѕРє РІСЃС–С… С‚СЏРіР°С‡С–РІ
TRUCKS = [
    "Р’Рђ1254Р•Рњ","BA1495EM","BA8603EM","Р’Рђ2387Р•РҐ","Р’Рђ5684РќРЎ",
    "BA8476EP","AM6937HE","Р’Рђ7289РќР†","Р’Рђ7286РќР†","Р’Рђ6675РќР†",
    "Р’Рђ6678РќР†","Р’Рђ2049РќРЎ","Р’Рђ9914Р•Рў","BA6468EP","Р’Рђ7287РќР†",
    "Р’Рђ8467РќР†","Р’Рђ5952РќРЎ","Р’Рђ7990Р•Рќ","Р’Рђ7954РђРћ","Р’Рђ1483Р•Рћ",
    "BA2187BK","Р’Рђ9244Р•Рќ","Р’Рђ8712Р•Р ","KA3566HE",
]

# РЎРїРёСЃРѕРє РІСЃС–С… С†РёСЃС‚РµСЂРЅ
TANKS = [
    "BA4872XO","AA3677XG","AA3622XG","BРҐ1209XF","BX0764XF",
    "BA0583XF","AA5938XG","Р’Рђ7565XF","Р’Рђ7566XF","Р’Рђ7716XF",
    "Р’Рђ7718XF","РђA2511XG","AA5942XG","BA0582XF","Р’Рђ7567XF",
    "Р’Рђ7719XF","РЎР•2735РҐР ","РЎР•2747РҐР ","Р’Рђ4694РҐРў","Р’Рђ4847РҐРћ",
    "BA6713XP","Р’Рђ4695РҐРў","BA4821РҐРћ","Р’Рђ5253РҐР ","Р’Рђ4954РҐРћ",
]

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# GOOGLE SHEETS
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
def get_spreadsheet():
    # РЎРїРѕС‡Р°С‚РєСѓ РїСЂРѕР±СѓС”РјРѕ РІР·СЏС‚Рё credentials Р·С– Р·РјС–РЅРЅРѕС— СЃРµСЂРµРґРѕРІРёС‰Р° (Р±РµР·РїРµС‡РЅС–С€Рµ)
    creds_json_env = os.environ.get("GOOGLE_CREDS_JSON")
    if creds_json_env:
        try:
            info = json.loads(creds_json_env)
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
            client = gspread.authorize(creds)
            return client.open_by_key(SHEET_ID)
        except Exception as e:
            logger.error(f"GOOGLE_CREDS_JSON parse error: {e}")

    # Р†РЅР°РєС€Рµ С€СѓРєР°С”РјРѕ JSON С„Р°Р№Р» Р· credentials
    candidates = [
        CREDS_FILE,
        "create-497113-eed86744057e.json..json",
        "create-497113-eed86744057e.json.json",
    ]
    try:
        for f in os.listdir("."):
            if f.endswith(".json") and "create-" in f.lower():
                if f not in candidates:
                    candidates.append(f)
    except Exception:
        pass

    creds_path = None
    for path in candidates:
        if os.path.exists(path):
            creds_path = path
            break

    if not creds_path:
        available = [f for f in os.listdir(".") if f.endswith(".json")] if os.path.exists(".") else []
        raise FileNotFoundError(
            f"Google credentials JSON РЅРµ Р·РЅР°Р№РґРµРЅРѕ. Р”РѕСЃС‚СѓРїРЅС– JSON: {available}"
        )

    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open_by_key(SHEET_ID)

def get_or_create_sheet(name, headers):
    ss = get_spreadsheet()
    try:
        ws = ss.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(name, rows=1000, cols=len(headers))
        ws.append_row(headers)
        ws.freeze(rows=1)
    return ws

def get_repairs_sheet():
    return get_or_create_sheet("Р РµРјРѕРЅС‚Рё", [
        "ID", "Р”Р°С‚Р° РїРѕРґР°С‡С–", "РўРёРї", "РњР°С€РёРЅР°", "РўРёРї РјР°С€РёРЅРё",
        "РћРїРёСЃ СЂРѕР±С–С‚ / Р—Р°РїС‡Р°СЃС‚РёРЅРё", "РЎСѓРјР°", "Р¤РѕСЂРјР° РѕРїР»Р°С‚Рё",
        "РЎРўРћ / РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє", "РќРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ", "РњРµРЅРµРґР¶РµСЂ",
        "РЎС‚Р°С‚СѓСЃ", "Р”Р°С‚Р° РїРѕРіРѕРґР¶РµРЅРЅСЏ", "Р”Р°С‚Р° РѕРїР»Р°С‚Рё"
    ])

def get_stock_sheet():
    return get_or_create_sheet("РЎРєР»Р°Рґ", [
        "РџРѕР·РёС†С–СЏ", "РћРґРёРЅРёС†СЏ", "РљС–Р»СЊРєС–СЃС‚СЊ", "Р¦С–РЅР° Р·Р° РѕРґРёРЅРёС†СЋ",
        "Р—Р°РіР°Р»СЊРЅР° РІР°СЂС‚С–СЃС‚СЊ", "Р”Р°С‚Р° РѕСЃС‚Р°РЅРЅСЊРѕРіРѕ РѕРЅРѕРІР»РµРЅРЅСЏ"
    ])

def get_movements_sheet():
    return get_or_create_sheet("Р СѓС… СЃРєР»Р°РґСѓ", [
        "ID", "Р”Р°С‚Р°", "РўРёРї", "РџРѕР·РёС†С–СЏ", "РљС–Р»СЊРєС–СЃС‚СЊ",
        "РњР°С€РёРЅР°", "РњРµРЅРµРґР¶РµСЂ", "РЎС‚Р°С‚СѓСЃ", "РџСЂРёРјС–С‚РєР°"
    ])

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РЎРљР›РђР” - Р›РћР“Р†РљРђ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
def get_stock_items():
    ws = get_stock_sheet()
    records = ws.get_all_records()
    return [r for r in records if r.get("РљС–Р»СЊРєС–СЃС‚СЊ", 0)]

def update_stock(position, unit, quantity_delta, price_per_unit=0):
    ws = get_stock_sheet()
    records = ws.get_all_records()
    for i, r in enumerate(records, start=2):
        if r.get("РџРѕР·РёС†С–СЏ", "").lower() == position.lower():
            new_qty = float(r.get("РљС–Р»СЊРєС–СЃС‚СЊ", 0)) + quantity_delta
            new_total = new_qty * float(r.get("Р¦С–РЅР° Р·Р° РѕРґРёРЅРёС†СЋ", price_per_unit) or price_per_unit)
            ws.update_cell(i, 3, round(new_qty, 3))
            ws.update_cell(i, 5, round(new_total, 2))
            ws.update_cell(i, 6, datetime.now().strftime("%d.%m.%Y"))
            return True
    if quantity_delta > 0:
        total = quantity_delta * price_per_unit
        ws.append_row([
            position, unit, round(quantity_delta, 3),
            round(price_per_unit, 2), round(total, 2),
            datetime.now().strftime("%d.%m.%Y")
        ])
    return True

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# AI Р РћР—РџР†Р—РќРђР’РђРќРќРЇ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def recognize_document(file_content: bytes, mime_type: str, is_pdf: bool = False) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)

    trucks_str = ", ".join(TRUCKS)
    tanks_str  = ", ".join(TANKS)

    prompt = f"""РўРё Р°РЅР°Р»С–Р·СѓС”С€ СЂР°С…СѓРЅРѕРє-С„Р°РєС‚СѓСЂСѓ РІС–Рґ РЎРўРћ Р°Р±Рѕ РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР° С‚РѕРІР°СЂС–РІ/РїРѕСЃР»СѓРі.

РЎРїРёСЃРѕРє С‚СЏРіР°С‡С–РІ РєРѕРјРїР°РЅС–С—: {trucks_str}
РЎРїРёСЃРѕРє С†РёСЃС‚РµСЂРЅ РєРѕРјРїР°РЅС–С—: {tanks_str}

Р’РђР–Р›РР’Рћ РїСЂРѕ РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР°:
- РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє (contractor) вЂ” С†Рµ РџР РћР”РђР’Р•Р¦Р¬, С‚РѕР№ С…С‚Рѕ Р’РРЎРўРђР’РР’ СЂР°С…СѓРЅРѕРє С– РЅР°РґР°РІ РїРѕСЃР»СѓРіРё/С‚РѕРІР°СЂРё.
- Р’ СЂР°С…СѓРЅРєСѓ С†Рµ РїРѕР»Рµ Р·Р°Р·РІРёС‡Р°Р№ РїРѕР·РЅР°С‡РµРЅРµ СЏРє "РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє", "РџСЂРѕРґР°РІРµС†СЊ", "Р’РёРєРѕРЅР°РІРµС†СЊ", "Р’РёРєРѕРЅР°РІРµС†СЊ РїРѕСЃР»СѓРі".
- РќР°С€Р° РєРѕРјРїР°РЅС–СЏ "РџРџ РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚" (Р°Р±Рѕ "РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚") С” РџРћРљРЈРџР¦Р•Рњ / Р—РђРњРћР’РќРРљРћРњ вЂ” С—С— РќР• С‚СЂРµР±Р° РІРєР°Р·СѓРІР°С‚Рё СЏРє РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР°!
- РЇРєС‰Рѕ РІ СЂР°С…СѓРЅРєСѓ РІРєР°Р·Р°РЅРѕ "РџРѕРєСѓРїРµС†СЊ: РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚" вЂ” С†Рµ РЅР°С€Р° РєРѕРјРїР°РЅС–СЏ, С–РіРЅРѕСЂСѓР№ С—С—.
- РЁСѓРєР°Р№ СЃР°РјРµ С‚РѕРіРѕ, РҐРўРћ Р’РРЎРўРђР’РР’ СЂР°С…СѓРЅРѕРє (Р·Р°Р·РІРёС‡Р°Р№ Р·РІРµСЂС…Сѓ РґРѕРєСѓРјРµРЅС‚Сѓ Р°Р±Рѕ РІ РїРѕР»С– "РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє").

Р’РёС‚СЏРіРЅРё Р· РґРѕРєСѓРјРµРЅС‚Р° РЅР°СЃС‚СѓРїРЅСѓ С–РЅС„РѕСЂРјР°С†С–СЋ Сѓ JSON:
- date: РґР°С‚Р° СЂР°С…СѓРЅРєСѓ (С„РѕСЂРјР°С‚ Р”Р”.РњРњ.Р Р Р Р )
- vehicle: РЅРѕРјРµСЂ РјР°С€РёРЅРё Р·С– СЃРїРёСЃРєСѓ РІРёС‰Рµ (СЏРєС‰Рѕ Р·РЅР°Р№РґРµРЅРѕ, С‚РѕС‡РЅРѕ СЏРє Сѓ СЃРїРёСЃРєСѓ)
- vehicle_type: "С‚СЏРіР°С‡" Р°Р±Рѕ "С†РёСЃС‚РµСЂРЅР°" (РІРёР·РЅР°С‡Рё Р·Р° РЅРѕРјРµСЂРѕРј)
- description: РѕРїРёСЃ СЂРѕР±С–С‚ Р°Р±Рѕ РїРµСЂРµР»С–Рє Р·Р°РїС‡Р°СЃС‚РёРЅ (РєРѕСЂРѕС‚РєРѕ)
- amount: СЃСѓРјР° (С‚С–Р»СЊРєРё С‡РёСЃР»Рѕ)
- payment_type: "Р±РµР·РЅР°Р»" Р°Р±Рѕ "РіРѕС‚С–РІРєР°"
- contractor: РЅР°Р·РІР° РџРћРЎРўРђР§РђР›Р¬РќРРљРђ/РџР РћР”РђР’Р¦РЇ (РќР• РїРѕРєСѓРїС†СЏ, РќР• РЅР°С€РѕС— РєРѕРјРїР°РЅС–С— РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚!)
- invoice_number: РЅРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ Р°Р±Рѕ РґРѕРєСѓРјРµРЅС‚Сѓ

РЇРєС‰Рѕ СЏРєРµСЃСЊ РїРѕР»Рµ РЅРµ Р·РЅР°Р№РґРµРЅРѕ вЂ” РІСЃС‚Р°РІ null.
Р’С–РґРїРѕРІС–РґР°Р№ РўР†Р›Р¬РљР JSON Р±РµР· Р¶РѕРґРЅРѕРіРѕ С–РЅС€РѕРіРѕ С‚РµРєСЃС‚Сѓ."""

    import base64
    b64 = base64.b64encode(file_content).decode()

    if is_pdf:
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
            {"type": "text", "text": prompt}
        ]
    else:
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": b64}},
            {"type": "text", "text": prompt}
        ]

    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        messages=[{"role": "user", "content": content}]
    )

    text = response.content[0].text.strip()
    text = re.sub(r'^```json\s*|\s*```$', '', text)
    return json.loads(text)


async def recognize_text_with_ai(text_content: str) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)

    trucks_str = ", ".join(TRUCKS)
    tanks_str  = ", ".join(TANKS)

    prompt = f"""РўРё Р°РЅР°Р»С–Р·СѓС”С€ С‚РµРєСЃС‚ СЂР°С…СѓРЅРєСѓ-С„Р°РєС‚СѓСЂРё, РІРёС‚СЏРіРЅСѓС‚РёР№ Р· Excel С‚Р°Р±Р»РёС†С–.

РЎРїРёСЃРѕРє С‚СЏРіР°С‡С–РІ РєРѕРјРїР°РЅС–С—: {trucks_str}
РЎРїРёСЃРѕРє С†РёСЃС‚РµСЂРЅ РєРѕРјРїР°РЅС–С—: {tanks_str}

Р’РђР–Р›РР’Рћ РїСЂРѕ РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР°:
- РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє (contractor) вЂ” С†Рµ РџР РћР”РђР’Р•Р¦Р¬, С‚РѕР№ С…С‚Рѕ Р’РРЎРўРђР’РР’ СЂР°С…СѓРЅРѕРє С– РЅР°РґР°РІ РїРѕСЃР»СѓРіРё/С‚РѕРІР°СЂРё.
- Р’ СЂР°С…СѓРЅРєСѓ С†Рµ РїРѕР»Рµ Р·Р°Р·РІРёС‡Р°Р№ РїРѕР·РЅР°С‡РµРЅРµ СЏРє "РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє", "РџСЂРѕРґР°РІРµС†СЊ", "Р’РёРєРѕРЅР°РІРµС†СЊ".
- РќР°С€Р° РєРѕРјРїР°РЅС–СЏ "РџРџ РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚" (Р°Р±Рѕ "РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚") С” РџРћРљРЈРџР¦Р•Рњ вЂ” С—С— РќР• С‚СЂРµР±Р° РІРєР°Р·СѓРІР°С‚Рё СЏРє РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР°!
- РЇРєС‰Рѕ РІ С‚РµРєСЃС‚С– С” "РџРѕРєСѓРїРµС†СЊ: РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚" вЂ” С†Рµ РЅР°С€Р° РєРѕРјРїР°РЅС–СЏ, С–РіРЅРѕСЂСѓР№ С—С—.

РўРµРєСЃС‚ СЂР°С…СѓРЅРєСѓ:
---
{text_content[:8000]}
---

Р’РёС‚СЏРіРЅРё РЅР°СЃС‚СѓРїРЅСѓ С–РЅС„РѕСЂРјР°С†С–СЋ Сѓ JSON:
- date: РґР°С‚Р° СЂР°С…СѓРЅРєСѓ (С„РѕСЂРјР°С‚ Р”Р”.РњРњ.Р Р Р Р )
- vehicle: РЅРѕРјРµСЂ РјР°С€РёРЅРё Р·С– СЃРїРёСЃРєСѓ РІРёС‰Рµ (СЏРєС‰Рѕ Р·РЅР°Р№РґРµРЅРѕ, С‚РѕС‡РЅРѕ СЏРє Сѓ СЃРїРёСЃРєСѓ)
- vehicle_type: "С‚СЏРіР°С‡" Р°Р±Рѕ "С†РёСЃС‚РµСЂРЅР°"
- description: РѕРїРёСЃ СЂРѕР±С–С‚ Р°Р±Рѕ РїРµСЂРµР»С–Рє РїРѕР·РёС†С–Р№ (РєРѕСЂРѕС‚РєРѕ, РґРѕ 100 СЃРёРјРІРѕР»С–РІ)
- amount: СЃСѓРјР° Р·Р°РіР°Р»СЊРЅР° (С‚С–Р»СЊРєРё С‡РёСЃР»Рѕ, Р±РµР· РїСЂРѕР±С–Р»С–РІ)
- payment_type: "Р±РµР·РЅР°Р»" Р°Р±Рѕ "РіРѕС‚С–РІРєР°"
- contractor: РЅР°Р·РІР° РџРћРЎРўРђР§РђР›Р¬РќРРљРђ (РќР• РўСЂР°РЅР·РёС‚-РўСЂР°СЃС‚!)
- invoice_number: РЅРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ

РЇРєС‰Рѕ СЏРєРµСЃСЊ РїРѕР»Рµ РЅРµ Р·РЅР°Р№РґРµРЅРѕ вЂ” РІСЃС‚Р°РІ null.
Р’С–РґРїРѕРІС–РґР°Р№ РўР†Р›Р¬РљР JSON Р±РµР· Р¶РѕРґРЅРѕРіРѕ С–РЅС€РѕРіРѕ С‚РµРєСЃС‚Сѓ."""

    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )

    result_text = response.content[0].text.strip()
    result_text = re.sub(r'^```json\s*|\s*```$', '', result_text)
    return json.loads(result_text)

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р“РћР›РћР’РќР• РњР•РќР®
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in MANAGER_IDS:
        await update.message.reply_text("в›” Р”РѕСЃС‚СѓРї Р·Р°Р±РѕСЂРѕРЅРµРЅРѕ.")
        return
    await show_main_menu(update, context)

async def show_main_menu(update, context):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("рџ”§ Р РµРјРѕРЅС‚", callback_data="menu_repair")],
        [InlineKeyboardButton("рџ“Ґ Р—Р°РєСѓРїРєР° РЅР° СЃРєР»Р°Рґ", callback_data="menu_purchase")],
        [InlineKeyboardButton("рџ“¤ РЎРїРёСЃР°РЅРЅСЏ Р·С– СЃРєР»Р°РґСѓ", callback_data="menu_writeoff")],
    ])
    text = "Р©Рѕ РґРѕРґР°С”РјРѕ?\n\nРћР±РµСЂС–С‚СЊ С‚РёРї РѕРїРµСЂР°С†С–С—:"
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=keyboard)
    else:
        await update.message.reply_text(text, reply_markup=keyboard)

async def new_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in MANAGER_IDS:
        return
    context.user_data.clear()
    await show_main_menu(update, context)

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РћР‘Р РћР‘РљРђ CALLBACK
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "menu_repair":
        context.user_data["op_type"] = "repair"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("рџ’µ Р“РѕС‚С–РІРєР° (РІСЂСѓС‡РЅСѓ)", callback_data="pay_cash")],
            [InlineKeyboardButton("рџЏ¦ Р‘РµР·РЅР°Р» (СЂР°С…СѓРЅРѕРє/С„РѕС‚Рѕ)", callback_data="pay_invoice")],
            [InlineKeyboardButton("в—ЂпёЏ РќР°Р·Р°Рґ", callback_data="back_main")],
        ])
        await query.edit_message_text("рџ”§ Р РµРјРѕРЅС‚\n\nРЎРїРѕСЃС–Р± РѕРїР»Р°С‚Рё:", reply_markup=keyboard)
        return

    if data == "menu_purchase":
        context.user_data["op_type"] = "purchase"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("рџ’µ Р“РѕС‚С–РІРєР° (РІСЂСѓС‡РЅСѓ)", callback_data="pay_cash")],
            [InlineKeyboardButton("рџЏ¦ Р‘РµР·РЅР°Р» (СЂР°С…СѓРЅРѕРє/С„РѕС‚Рѕ)", callback_data="pay_invoice")],
            [InlineKeyboardButton("в—ЂпёЏ РќР°Р·Р°Рґ", callback_data="back_main")],
        ])
        await query.edit_message_text("рџ“Ґ Р—Р°РєСѓРїРєР° РЅР° СЃРєР»Р°Рґ\n\nРЎРїРѕСЃС–Р± РѕРїР»Р°С‚Рё:", reply_markup=keyboard)
        return

    if data == "menu_writeoff":
        context.user_data["op_type"] = "writeoff"
        await show_stock_selection(query, context)
        return

    if data == "back_main":
        context.user_data.clear()
        await show_main_menu(update, context)
        return

    if data == "pay_cash":
        context.user_data["payment"] = "РіРѕС‚С–РІРєР°"
        op = context.user_data.get("op_type")
        if op == "repair":
            await ask_vehicle(query, context)
        else:
            context.user_data["step"] = "manual_purchase"
            await query.edit_message_text(
                "рџ“Ґ Р’РІРµРґС–С‚СЊ РґР°РЅС– Р·Р°РєСѓРїРєРё РІ РѕРґРёРЅ СЂСЏРґРѕРє:\n\n"
                "<code>РќР°Р·РІР° РїРѕР·РёС†С–С—, РєС–Р»СЊРєС–СЃС‚СЊ, РѕРґРёРЅРёС†СЏ, СЃСѓРјР°, РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє</code>\n\n"
                "РџСЂРёРєР»Р°Рґ:\n"
                "<code>РњР°СЃР»Рѕ РјРѕС‚РѕСЂРЅРµ 5W-40, 200, Р», 50000, РЈРєСЂРЅР°С„С‚Р°</code>",
                parse_mode="HTML"
            )
        return

    if data == "pay_invoice":
        context.user_data["payment"] = "Р±РµР·РЅР°Р»"
        context.user_data["step"] = "awaiting_file"
        await query.edit_message_text(
            "рџ“Ћ РќР°РґС–С€Р»С–С‚СЊ СЂР°С…СѓРЅРѕРє:\n\n"
            "вЂў PDF С„Р°Р№Р»\n"
            "вЂў Р¤РѕС‚Рѕ СЂР°С…СѓРЅРєСѓ\n"
            "вЂў РЎРєР°РЅ РґРѕРєСѓРјРµРЅС‚Сѓ\n\n"
            "AI СЂРѕР·РїС–Р·РЅР°С” РґР°РЅС– Р°РІС‚РѕРјР°С‚РёС‡РЅРѕ."
        )
        return

    if data.startswith("truck_"):
        vehicle = data[6:]
        context.user_data["vehicle"] = vehicle
        context.user_data["vehicle_type"] = "С‚СЏРіР°С‡"
        await ask_manual_repair(query, context)
        return

    if data.startswith("tank_"):
        vehicle = data[5:]
        context.user_data["vehicle"] = vehicle
        context.user_data["vehicle_type"] = "С†РёСЃС‚РµСЂРЅР°"
        await ask_manual_repair(query, context)
        return

    if data == "vehicle_page_trucks":
        await show_vehicle_selection(query, context, "trucks", 0)
        return

    if data == "vehicle_page_tanks":
        await show_vehicle_selection(query, context, "tanks", 0)
        return

    if data.startswith("vpage_"):
        _, vtype, page = data.split("_")
        await show_vehicle_selection(query, context, vtype, int(page))
        return

    if data.startswith("stock_item_"):
        item_idx = int(data[11:])
        items = get_stock_items()
        if item_idx < len(items):
            item = items[item_idx]
            context.user_data["stock_item"] = item
            context.user_data["step"] = "writeoff_qty"
            await query.edit_message_text(
                f"рџ“¤ РЎРїРёСЃР°РЅРЅСЏ: <b>{item['РџРѕР·РёС†С–СЏ']}</b>\n"
                f"РќР° СЃРєР»Р°РґС–: {item['РљС–Р»СЊРєС–СЃС‚СЊ']} {item['РћРґРёРЅРёС†СЏ']}\n\n"
                f"Р’РІРµРґС–С‚СЊ РєС–Р»СЊРєС–СЃС‚СЊ РґР»СЏ СЃРїРёСЃР°РЅРЅСЏ:",
                parse_mode="HTML"
            )
        return

    if data == "confirm_data":
        await submit_for_approval(query, context)
        return

    if data == "edit_data":
        await show_edit_menu(query, context)
        return

    if data.startswith("edit_field_"):
        field = data[11:]
        context.user_data["editing_field"] = field
        context.user_data["step"] = "editing_field"
        field_labels = {
            "vehicle": "рџљ— РјР°С€РёРЅСѓ (РІРІРµРґС–С‚СЊ РЅРѕРјРµСЂ, РЅР°РїСЂРёРєР»Р°Рґ BA2187BK)",
            "description": "рџ“ќ РѕРїРёСЃ СЂРѕР±С–С‚ / Р·Р°РїС‡Р°СЃС‚РёРЅ",
            "amount": "рџ’° СЃСѓРјСѓ (С‚С–Р»СЊРєРё С‡РёСЃР»Рѕ)",
            "contractor": "рџЏЄ РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРєР° / РЎРўРћ",
            "date": "рџ“… РґР°С‚Сѓ (С„РѕСЂРјР°С‚ Р”Р”.РњРњ.Р Р Р Р )",
            "invoice": "рџ”ў РЅРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ",
        }
        await query.edit_message_text(
            f"Р’РІРµРґС–С‚СЊ РЅРѕРІРµ Р·РЅР°С‡РµРЅРЅСЏ РґР»СЏ РїРѕР»СЏ {field_labels.get(field, field)}:"
        )
        return

    if data == "confirm_vehicle":
        await ask_manual_repair(query, context)
        return

    if data.startswith("director_approve_"):
        record_id = int(data[17:])
        await director_approve(query, context, record_id)
        return

    if data.startswith("director_reject_"):
        record_id = int(data[16:])
        await director_reject(query, context, record_id)
        return

    if data.startswith("accountant_paid_"):
        record_id = int(data[16:])
        await accountant_paid(query, context, record_id)
        return

    if data.startswith("wo_truck_"):
        vehicle = data[9:]
        context.user_data["vehicle"] = vehicle
        context.user_data["vehicle_type"] = "С‚СЏРіР°С‡"
        await submit_writeoff(query, context)
        return

    if data.startswith("wo_tank_"):
        vehicle = data[8:]
        context.user_data["vehicle"] = vehicle
        context.user_data["vehicle_type"] = "С†РёСЃС‚РµСЂРЅР°"
        await submit_writeoff(query, context)
        return

    if data == "wo_page_trucks":
        await show_writeoff_vehicle(query, context, "trucks", 0)
        return

    if data == "wo_page_tanks":
        await show_writeoff_vehicle(query, context, "tanks", 0)
        return

    if data.startswith("wopage_"):
        _, vtype, page = data.split("_")
        await show_writeoff_vehicle(query, context, vtype, int(page))
        return


# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р’РР‘Р†Р  РњРђРЁРРќР
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def ask_vehicle(query, context):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("рџљ› РўСЏРіР°С‡", callback_data="vehicle_page_trucks")],
        [InlineKeyboardButton("рџ›ў Р¦РёСЃС‚РµСЂРЅР°", callback_data="vehicle_page_tanks")],
        [InlineKeyboardButton("в—ЂпёЏ РќР°Р·Р°Рґ", callback_data="back_main")],
    ])
    op = context.user_data.get("op_type", "")
    op_label = "рџ”§ Р РµРјРѕРЅС‚" if op == "repair" else "рџ“Ґ Р—Р°РєСѓРїРєР°"
    await query.edit_message_text(
        f"{op_label}\n\nРћР±РµСЂС–С‚СЊ С‚РёРї С‚СЂР°РЅСЃРїРѕСЂС‚Сѓ:",
        reply_markup=keyboard
    )

async def show_vehicle_selection(query, context, vtype, page):
    items = TRUCKS if vtype == "trucks" else TANKS
    prefix = "truck_" if vtype == "trucks" else "tank_"
    per_page = 8
    start = page * per_page
    end = min(start + per_page, len(items))
    chunk = items[start:end]

    rows = []
    for i in range(0, len(chunk), 2):
        row = [InlineKeyboardButton(chunk[i], callback_data=f"{prefix}{chunk[i]}")]
        if i + 1 < len(chunk):
            row.append(InlineKeyboardButton(chunk[i+1], callback_data=f"{prefix}{chunk[i+1]}"))
        rows.append(row)

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("в—ЂпёЏ", callback_data=f"vpage_{vtype}_{page-1}"))
    if end < len(items):
        nav.append(InlineKeyboardButton("в–¶пёЏ", callback_data=f"vpage_{vtype}_{page+1}"))
    if nav:
        rows.append(nav)

    other = "tanks" if vtype == "trucks" else "trucks"
    other_label = "рџ›ў Р¦РёСЃС‚РµСЂРЅРё" if vtype == "trucks" else "рџљ› РўСЏРіР°С‡С–"
    rows.append([InlineKeyboardButton(f"в†©пёЏ {other_label}", callback_data=f"vehicle_page_{other}")])

    label = "РўСЏРіР°С‡С–" if vtype == "trucks" else "Р¦РёСЃС‚РµСЂРЅРё"
    await query.edit_message_text(
        f"РћР±РµСЂС–С‚СЊ {label} ({start+1}-{end} Р· {len(items)}):",
        reply_markup=InlineKeyboardMarkup(rows)
    )

async def ask_manual_repair(query, context):
    vehicle = context.user_data.get("vehicle", "")
    vtype   = context.user_data.get("vehicle_type", "")
    payment = context.user_data.get("payment", "РіРѕС‚С–РІРєР°")
    op      = context.user_data.get("op_type", "repair")

    context.user_data["step"] = "manual_input"

    if op == "repair":
        await query.edit_message_text(
            f"рџ”§ Р РµРјРѕРЅС‚ вЂ” {vtype} {vehicle}\n"
            f"РћРїР»Р°С‚Р°: {payment}\n\n"
            f"Р’РІРµРґС–С‚СЊ РґР°РЅС– РІ РѕРґРЅРѕРјСѓ РїРѕРІС–РґРѕРјР»РµРЅРЅС–:\n\n"
            f"<code>РћРїРёСЃ СЂРѕР±С–С‚, СЃСѓРјР°, РЎРўРћ, РЅРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ (СЏРєС‰Рѕ С”)</code>\n\n"
            f"РџСЂРёРєР»Р°Рґ:\n"
            f"<code>Р—Р°РјС–РЅР° РіР°Р»СЊРјС–РІРЅРёС… РєРѕР»РѕРґРѕРє, 8500, РЎРўРћ РђРІС‚РѕРјР°Р№СЃС‚РµСЂ, СЂР°С….128</code>",
            parse_mode="HTML"
        )
    else:
        await query.edit_message_text(
            f"рџ“Ґ Р—Р°РєСѓРїРєР° вЂ” {vtype} {vehicle}\n"
            f"РћРїР»Р°С‚Р°: {payment}\n\n"
            f"Р’РІРµРґС–С‚СЊ РґР°РЅС–:\n\n"
            f"<code>РќР°Р·РІР°, РєС–Р»СЊРєС–СЃС‚СЊ, РѕРґРёРЅРёС†СЏ, СЃСѓРјР°, РїРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє</code>\n\n"
            f"РџСЂРёРєР»Р°Рґ:\n"
            f"<code>Р¤С–Р»СЊС‚СЂ РѕР»РёРІРё, 12, С€С‚, 3600, РђРІС‚РѕР·Р°РїС‡Р°СЃС‚РёРЅРё</code>",
            parse_mode="HTML"
        )

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р’РР‘Р†Р  РџРћР—РР¦Р†Р‡ Р—Р† РЎРљР›РђР”РЈ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def show_stock_selection(query, context):
    try:
        items = get_stock_items()
    except Exception:
        items = []

    if not items:
        await query.edit_message_text(
            "рџ“¦ РЎРєР»Р°Рґ РїРѕСЂРѕР¶РЅС–Р№.\n\n"
            "РЎРїРѕС‡Р°С‚РєСѓ Р·СЂРѕР±С–С‚СЊ Р·Р°РєСѓРїРєСѓ С‡РµСЂРµР· рџ“Ґ Р—Р°РєСѓРїРєР° РЅР° СЃРєР»Р°Рґ."
        )
        return

    rows = []
    for i, item in enumerate(items[:10]):
        label = f"{item['РџРѕР·РёС†С–СЏ']} вЂ” {item['РљС–Р»СЊРєС–СЃС‚СЊ']} {item['РћРґРёРЅРёС†СЏ']}"
        rows.append([InlineKeyboardButton(label, callback_data=f"stock_item_{i}")])
    rows.append([InlineKeyboardButton("в—ЂпёЏ РќР°Р·Р°Рґ", callback_data="back_main")])

    await query.edit_message_text(
        "рџ“¤ РЎРїРёСЃР°РЅРЅСЏ Р·С– СЃРєР»Р°РґСѓ\n\nРћР±РµСЂС–С‚СЊ РїРѕР·РёС†С–СЋ:",
        reply_markup=InlineKeyboardMarkup(rows)
    )

async def show_writeoff_vehicle(query, context, vtype, page):
    items = TRUCKS if vtype == "trucks" else TANKS
    prefix = "wo_truck_" if vtype == "trucks" else "wo_tank_"
    per_page = 8
    start = page * per_page
    end = min(start + per_page, len(items))
    chunk = items[start:end]

    rows = []
    for i in range(0, len(chunk), 2):
        row = [InlineKeyboardButton(chunk[i], callback_data=f"{prefix}{chunk[i]}")]
        if i + 1 < len(chunk):
            row.append(InlineKeyboardButton(chunk[i+1], callback_data=f"{prefix}{chunk[i+1]}"))
        rows.append(row)

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("в—ЂпёЏ", callback_data=f"wopage_{vtype}_{page-1}"))
    if end < len(items):
        nav.append(InlineKeyboardButton("в–¶пёЏ", callback_data=f"wopage_{vtype}_{page+1}"))
    if nav:
        rows.append(nav)

    other = "tanks" if vtype == "trucks" else "trucks"
    other_label = "рџ›ў Р¦РёСЃС‚РµСЂРЅРё" if vtype == "trucks" else "рџљ› РўСЏРіР°С‡С–"
    rows.append([InlineKeyboardButton(f"в†©пёЏ {other_label}", callback_data=f"wo_page_{other}")])

    label = "РўСЏРіР°С‡С–" if vtype == "trucks" else "Р¦РёСЃС‚РµСЂРЅРё"
    item = context.user_data.get("stock_item", {})
    await query.edit_message_text(
        f"рџ“¤ РЎРїРёСЃР°РЅРЅСЏ: {item.get('РџРѕР·РёС†С–СЏ','')}\n"
        f"РљС–Р»СЊРєС–СЃС‚СЊ: {context.user_data.get('writeoff_qty','')} {item.get('РћРґРёРЅРёС†СЏ','')}\n\n"
        f"РћР±РµСЂС–С‚СЊ {label} ({start+1}-{end} Р· {len(items)}):",
        reply_markup=InlineKeyboardMarkup(rows)
    )

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РћР‘Р РћР‘РљРђ РўР•РљРЎРўРћР’РРҐ РџРћР’Р†Р”РћРњР›Р•РќР¬
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in MANAGER_IDS:
        return
    if update.message.text.startswith("/"):
        return

    step = context.user_data.get("step")

    # Р РµРґР°РіСѓРІР°РЅРЅСЏ РєРѕРЅРєСЂРµС‚РЅРѕРіРѕ РїРѕР»СЏ
    if step == "editing_field":
        field = context.user_data.get("editing_field")
        new_value = update.message.text.strip()

        if field == "vehicle":
            found = False
            for v in TRUCKS + TANKS:
                if v.upper() == new_value.upper():
                    context.user_data["vehicle"] = v
                    context.user_data["vehicle_type"] = "С‚СЏРіР°С‡" if v in TRUCKS else "С†РёСЃС‚РµСЂРЅР°"
                    found = True
                    break
            if not found:
                context.user_data["vehicle"] = new_value
        else:
            field_map = {
                "description": "description",
                "amount": "amount",
                "contractor": "contractor",
                "date": "date",
                "invoice": "invoice",
            }
            key = field_map.get(field, field)
            if field == "amount":
                new_value = new_value.replace(" ", "").replace(",", ".")
            context.user_data[key] = new_value

        context.user_data.pop("editing_field", None)
        context.user_data["step"] = "confirming"
        await show_confirmation(update.message, context)
        return

    # РљС–Р»СЊРєС–СЃС‚СЊ РґР»СЏ СЃРїРёСЃР°РЅРЅСЏ
    if step == "writeoff_qty":
        text = update.message.text.strip().replace(",", ".")
        try:
            qty = float(text)
        except ValueError:
            await update.message.reply_text("Р’РІРµРґС–С‚СЊ С‡РёСЃР»Рѕ (РєС–Р»СЊРєС–СЃС‚СЊ). РќР°РїСЂРёРєР»Р°Рґ: 20 Р°Р±Рѕ 20.5")
            return
        item = context.user_data.get("stock_item", {})
        available = float(item.get("РљС–Р»СЊРєС–СЃС‚СЊ", 0))
        if qty > available:
            await update.message.reply_text(
                f"вќ—пёЏ РќР° СЃРєР»Р°РґС– С‚С–Р»СЊРєРё {available} {item.get('РћРґРёРЅРёС†СЏ','')}.\n"
                f"Р’РІРµРґС–С‚СЊ РјРµРЅС€Рµ Р°Р±Рѕ СЂС–РІРЅРѕ {available}:"
            )
            return
        context.user_data["writeoff_qty"] = qty
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("рџљ› РўСЏРіР°С‡", callback_data="wo_page_trucks")],
            [InlineKeyboardButton("рџ›ў Р¦РёСЃС‚РµСЂРЅР°", callback_data="wo_page_tanks")],
        ])
        await update.message.reply_text(
            f"рџ“¤ {item.get('РџРѕР·РёС†С–СЏ','')} вЂ” {qty} {item.get('РћРґРёРЅРёС†СЏ','')}\n\n"
            f"РќР° СЏРєСѓ РјР°С€РёРЅСѓ СЃРїРёСЃСѓС”РјРѕ?",
            reply_markup=keyboard
        )
        return

    if step == "manual_input":
        parts = [p.strip() for p in update.message.text.split(",")]
        op = context.user_data.get("op_type")

        if op == "repair":
            description = parts[0] if len(parts) > 0 else ""
            amount = parts[1] if len(parts) > 1 else ""
            contractor = parts[2] if len(parts) > 2 else ""
            invoice = parts[3] if len(parts) > 3 else ""

            context.user_data.update({
                "description": description,
                "amount": amount.replace(" ", ""),
                "contractor": contractor,
                "invoice": invoice,
                "date": datetime.now().strftime("%d.%m.%Y"),
            })
        else:
            name = parts[0] if len(parts) > 0 else ""
            qty  = parts[1] if len(parts) > 1 else ""
            unit = parts[2] if len(parts) > 2 else ""
            amount = parts[3] if len(parts) > 3 else ""
            contractor = parts[4] if len(parts) > 4 else ""

            context.user_data.update({
                "description": f"{name} {qty} {unit}",
                "stock_name": name,
                "stock_qty": qty,
                "stock_unit": unit,
                "amount": amount.replace(" ", ""),
                "contractor": contractor,
                "invoice": "",
                "date": datetime.now().strftime("%d.%m.%Y"),
            })

        context.user_data["step"] = "confirming"
        await show_confirmation(update.message, context)
        return

    if step == "manual_purchase":
        parts = [p.strip() for p in update.message.text.split(",")]
        name = parts[0] if len(parts) > 0 else ""
        qty  = parts[1] if len(parts) > 1 else ""
        unit = parts[2] if len(parts) > 2 else ""
        amount = parts[3] if len(parts) > 3 else ""
        contractor = parts[4] if len(parts) > 4 else ""

        context.user_data.update({
            "description": f"{name} {qty} {unit}",
            "stock_name": name, "stock_qty": qty, "stock_unit": unit,
            "amount": amount.replace(" ", ""), "contractor": contractor,
            "invoice": "", "date": datetime.now().strftime("%d.%m.%Y"),
            "vehicle": "", "vehicle_type": "",
            "step": "confirming",
        })
        await show_confirmation(update.message, context)
        return


async def show_edit_menu(msg_obj, context):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("рџљ— РњР°С€РёРЅР°", callback_data="edit_field_vehicle")],
        [InlineKeyboardButton("рџ“ќ РћРїРёСЃ", callback_data="edit_field_description")],
        [InlineKeyboardButton("рџ’° РЎСѓРјР°", callback_data="edit_field_amount")],
        [InlineKeyboardButton("рџЏЄ РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє/РЎРўРћ", callback_data="edit_field_contractor")],
        [InlineKeyboardButton("рџ“… Р”Р°С‚Р°", callback_data="edit_field_date")],
        [InlineKeyboardButton("рџ”ў РќРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ", callback_data="edit_field_invoice")],
        [InlineKeyboardButton("вњ… Р“РѕС‚РѕРІРѕ, РІС–РґРїСЂР°РІРёС‚Рё", callback_data="confirm_data")],
    ])
    text = "вњЏпёЏ <b>Р©Рѕ РІРёРїСЂР°РІРёС‚Рё?</b>\n\nРћР±РµСЂС–С‚СЊ РїРѕР»Рµ РґР»СЏ СЂРµРґР°РіСѓРІР°РЅРЅСЏ:"

    if hasattr(msg_obj, 'reply_text'):
        await msg_obj.reply_text(text, parse_mode="HTML", reply_markup=keyboard)
    else:
        try:
            await msg_obj.edit_message_text(text, parse_mode="HTML", reply_markup=keyboard)
        except Exception:
            await msg_obj.message.reply_text(text, parse_mode="HTML", reply_markup=keyboard)


async def show_confirmation(msg_obj, context):
    op      = context.user_data.get("op_type", "")
    vehicle = context.user_data.get("vehicle", "вЂ”")
    vtype   = context.user_data.get("vehicle_type", "")
    desc    = context.user_data.get("description", "")
    amount  = context.user_data.get("amount", "")
    pay     = context.user_data.get("payment", "")
    contr   = context.user_data.get("contractor", "")
    invoice = context.user_data.get("invoice", "")
    date    = context.user_data.get("date", "")

    op_labels = {"repair": "рџ”§ Р РµРјРѕРЅС‚", "purchase": "рџ“Ґ Р—Р°РєСѓРїРєР°", "writeoff": "рџ“¤ РЎРїРёСЃР°РЅРЅСЏ"}
    op_label = op_labels.get(op, op)

    text = (
        f"<b>РџРµСЂРµРІС–СЂС‚Рµ РґР°РЅС–:</b>\n\n"
        f"рџ“‹ РўРёРї: {op_label}\n"
        f"рџљ— РњР°С€РёРЅР°: {vtype} {vehicle}\n"
        f"рџ“… Р”Р°С‚Р°: {date}\n"
        f"рџ“ќ РћРїРёСЃ: {desc}\n"
        f"рџ’° РЎСѓРјР°: {amount} РіСЂРЅ\n"
        f"рџ’і РћРїР»Р°С‚Р°: {pay}\n"
        f"рџЏЄ РџРѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє/РЎРўРћ: {contr}\n"
    )
    if invoice:
        text += f"рџ”ў Р Р°С…СѓРЅРѕРє: {invoice}\n"
    text += "\nР’СЃРµ РІС–СЂРЅРѕ?"

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("вњ… РўР°Рє, РІС–РґРїСЂР°РІРёС‚Рё", callback_data="confirm_data"),
            InlineKeyboardButton("вњЏпёЏ Р’РёРїСЂР°РІРёС‚Рё", callback_data="edit_data"),
        ]
    ])

    if hasattr(msg_obj, 'reply_text'):
        await msg_obj.reply_text(text, parse_mode="HTML", reply_markup=keyboard)
    else:
        try:
            await msg_obj.edit_message_text(text, parse_mode="HTML", reply_markup=keyboard)
        except Exception:
            await msg_obj.message.reply_text(text, parse_mode="HTML", reply_markup=keyboard)

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РћР‘Р РћР‘РљРђ Р¤РђР™Р›Р†Р’ (PDF, С„РѕС‚Рѕ, Excel)
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
def extract_excel_text(file_bytes: bytes) -> str:
    """Р’РёС‚СЏРіСѓС” РІРµСЃСЊ С‚РµРєСЃС‚ Р· Excel С„Р°Р№Р»Сѓ."""
    import io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    all_text.append(" | ".join(row_text))
        return "\n".join(all_text)
    except Exception as e:
        logger.error(f"openpyxl error: {e}")
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            all_text = []
            for sheet_idx in range(wb_xls.nsheets):
                ws_xls = wb_xls.sheet_by_index(sheet_idx)
                for r in range(ws_xls.nrows):
                    row_text = []
                    for c in range(ws_xls.ncols):
                        v = ws_xls.cell_value(r, c)
                        if v:
                            row_text.append(str(v))
                    if row_text:
                        all_text.append(" | ".join(row_text))
            return "\n".join(all_text)
        except Exception as e2:
            logger.error(f"xlrd error: {e2}")
            return ""


async def parse_excel_invoice(file_bytes: bytes) -> dict:
    """РџР°СЂСЃРёС‚СЊ Excel СЂР°С…СѓРЅРѕРє: РІРёС‚СЏРіСѓС” С‚РµРєСЃС‚ С– СЂРѕР·РїС–Р·РЅР°С” С‡РµСЂРµР· AI."""
    text = extract_excel_text(file_bytes)
    if not text:
        return {}

    try:
        return await recognize_text_with_ai(text)
    except Exception as e:
        logger.error(f"Excel AI recognition error: {e}")
        return extract_invoice_data_from_text(text)


def extract_invoice_data_from_text(text: str) -> dict:
    """Р РµР·РµСЂРІРЅРёР№ regex-РїР°СЂСЃРµСЂ."""
    result = {}

    amount_match = re.search(r'[РЈСѓ]СЃСЊРѕРіРѕ\s+Р·\s+РџР”Р’[:\s]+([0-9\s,\.]+)', text)
    if not amount_match:
        amount_match = re.search(r'[Р’РІ]СЃСЊРѕРіРѕ[:\s]+([0-9\s,\.]+)\s*РіСЂРЅ', text)
    if not amount_match:
        numbers = re.findall(r'(\d{3,7}(?:[,\.]\d{2})?)', text)
        if numbers:
            amounts = [float(n.replace(',', '.').replace(' ', '')) for n in numbers]
            result["amount"] = int(max(amounts))
    else:
        amt = amount_match.group(1).strip().replace(' ', '').replace(',', '.')
        try:
            result["amount"] = int(float(amt))
        except Exception:
            pass

    inv_match = re.search(r'[Р СЂ]Р°С…СѓРЅРѕРє[^\d]*(?:в„–\s*)?(\d+)', text)
    if inv_match:
        result["invoice"] = inv_match.group(1)

    date_match = re.search(r'(\d{1,2})[.\s]+([Р°-СЏРђ-РЇС‘РЃС–Р†С—Р‡С”Р„]+)\s+(\d{4})', text)
    if date_match:
        months = {"СЃС–С‡РЅСЏ":1,"Р»СЋС‚РѕРіРѕ":2,"Р±РµСЂРµР·РЅСЏ":3,"РєРІС–С‚РЅСЏ":4,"С‚СЂР°РІРЅСЏ":5,"С‡РµСЂРІРЅСЏ":6,
                  "Р»РёРїРЅСЏ":7,"СЃРµСЂРїРЅСЏ":8,"РІРµСЂРµСЃРЅСЏ":9,"Р¶РѕРІС‚РЅСЏ":10,"Р»РёСЃС‚РѕРїР°РґР°":11,"РіСЂСѓРґРЅСЏ":12}
        m = months.get(date_match.group(2).lower())
        if m:
            result["date"] = f"{int(date_match.group(1)):02d}.{m:02d}.{date_match.group(3)}"
    if not result.get("date"):
        date_match2 = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', text)
        if date_match2:
            result["date"] = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"

    sup_match = re.search(r"[РџРї]РѕСЃС‚Р°С‡Р°Р»СЊРЅРёРє[^\n]{0,5}([Рђ-РЇТђР„Р†Р‡С‘A-Z][^\n]{3,50})", text)
    if sup_match:
        result["contractor"] = sup_match.group(1).strip()

    desc_patterns = [r"Р—Р°РїСЂР°РІРєР°[^\n]{0,50}", r"Р РµРјРѕРЅС‚[^\n]{0,50}", r"[Р—Р·]Р°РјС–РЅР°[^\n]{0,50}", r"РўРћ[^\n]{0,30}"]
    for pat in desc_patterns:
        dm = re.search(pat, text)
        if dm:
            result["description"] = dm.group(0).strip()[:100]
            break

    for vehicle in TRUCKS + TANKS:
        if vehicle.upper() in text.upper():
            result["vehicle"] = vehicle
            result["vehicle_type"] = "С‚СЏРіР°С‡" if vehicle in TRUCKS else "С†РёСЃС‚РµСЂРЅР°"
            break

    return result


async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in MANAGER_IDS:
        return
    if context.user_data.get("step") != "awaiting_file":
        return

    await update.message.reply_text("рџ¤– AI СЂРѕР·РїС–Р·РЅР°С” РґРѕРєСѓРјРµРЅС‚, Р·Р°С‡РµРєР°Р№С‚Рµ...")

    try:
        file_bytes = None
        mime = "application/pdf"
        is_pdf = False
        is_excel = False

        if update.message.document:
            file = await context.bot.get_file(update.message.document.file_id)
            file_bytes = bytes(await file.download_as_bytearray())
            mime = update.message.document.mime_type or "application/pdf"
            is_pdf = "pdf" in mime.lower()
            is_excel = any(x in mime.lower() for x in ["excel", "spreadsheet", "xls"]) or \
                       update.message.document.file_name.endswith((".xls", ".xlsx"))
            context.user_data["file_id"] = update.message.document.file_id
            context.user_data["file_type"] = "document"
        elif update.message.photo:
            file = await context.bot.get_file(update.message.photo[-1].file_id)
            file_bytes = bytes(await file.download_as_bytearray())
            mime = "image/jpeg"
            context.user_data["file_id"] = update.message.photo[-1].file_id
            context.user_data["file_type"] = "photo"
        else:
            await update.message.reply_text("РќР°РґС–С€Р»С–С‚СЊ PDF, С„РѕС‚Рѕ Р°Р±Рѕ Excel С„Р°Р№Р» СЂР°С…СѓРЅРєСѓ.")
            return

        data = {}
        if is_excel:
            try:
                data = await parse_excel_invoice(file_bytes)
            except Exception as e:
                logger.error(f"Excel parse error: {e}")
        else:
            try:
                data = await recognize_document(file_bytes, mime, is_pdf)
            except Exception as e:
                logger.error(f"AI recognition error: {e}")

        context.user_data.update({
            "date":         data.get("date") or datetime.now().strftime("%d.%m.%Y"),
            "vehicle":      data.get("vehicle") or context.user_data.get("vehicle", ""),
            "vehicle_type": data.get("vehicle_type") or context.user_data.get("vehicle_type", ""),
            "description":  data.get("description") or "",
            "amount":       str(data.get("amount") or ""),
            "contractor":   data.get("contractor") or "",
            "invoice":      data.get("invoice") or "",
            "step":         "confirming",
        })

        await show_confirmation(update.message, context)

    except Exception as e:
        logger.error(f"handle_file error: {e}")
        await update.message.reply_text(
            "вќ—пёЏ РќРµ РІРґР°Р»РѕСЃСЊ СЂРѕР·РїС–Р·РЅР°С‚Рё РґРѕРєСѓРјРµРЅС‚.\n\n"
            "Р’РІРµРґС–С‚СЊ РґР°РЅС– РІСЂСѓС‡РЅСѓ:\n"
            "<code>РћРїРёСЃ, СЃСѓРјР°, РЎРўРћ, РЅРѕРјРµСЂ СЂР°С…СѓРЅРєСѓ</code>",
            parse_mode="HTML"
        )
        context.user_data["step"] = "manual_input"

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р’Р†Р”РџР РђР’РљРђ РќРђ РџРћР“РћР”Р–Р•РќРќРЇ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def submit_for_approval(query, context):
    try:
        ws = get_repairs_sheet()
        all_rows = ws.get_all_values()
        record_id = len(all_rows)

        op      = context.user_data.get("op_type", "")
        vehicle = context.user_data.get("vehicle", "")
        vtype   = context.user_data.get("vehicle_type", "")
        desc    = context.user_data.get("description", "")
        amount  = context.user_data.get("amount", "")
        pay     = context.user_data.get("payment", "РіРѕС‚С–РІРєР°")
        contr   = context.user_data.get("contractor", "")
        invoice = context.user_data.get("invoice", "")
        date    = context.user_data.get("date", "")
        manager = MANAGER_NAMES.get(query.from_user.id, query.from_user.first_name)

        op_labels = {"repair": "Р РµРјРѕРЅС‚", "purchase": "Р—Р°РєСѓРїРєР°", "writeoff": "РЎРїРёСЃР°РЅРЅСЏ"}
        op_label = op_labels.get(op, op)

        ws.append_row([
            record_id,
            datetime.now().strftime("%d.%m.%Y %H:%M"),
            op_label,
            vehicle,
            vtype,
            desc,
            amount,
            pay,
            contr,
            invoice,
            manager,
            "РќР° РїРѕРіРѕРґР¶РµРЅРЅС–",
            "", ""
        ])

        context.application.bot_data[f"repair_{record_id}"] = dict(context.user_data)
        context.application.bot_data[f"repair_{record_id}"]["record_id"] = record_id
        context.application.bot_data[f"repair_{record_id}"]["manager_id"] = query.from_user.id

        await query.edit_message_text(
            f"вњ… Р—Р°СЏРІРєСѓ #{record_id} РІС–РґРїСЂР°РІР»РµРЅРѕ РЅР° РїРѕРіРѕРґР¶РµРЅРЅСЏ РґРёСЂРµРєС‚РѕСЂСѓ.\n\n"
            f"РћС‡С–РєСѓР№С‚Рµ СЂС–С€РµРЅРЅСЏ."
        )

        director_text = (
            f"рџ“‹ <b>РќРѕРІР° Р·Р°СЏРІРєР° #{record_id}</b>\n\n"
            f"рџ‘¤ РњРµРЅРµРґР¶РµСЂ: {manager}\n"
            f"рџ“‹ РўРёРї: {op_label}\n"
            f"рџљ— РњР°С€РёРЅР°: {vtype} {vehicle}\n"
            f"рџ“… Р”Р°С‚Р°: {date}\n"
            f"рџ“ќ РћРїРёСЃ: {desc}\n"
            f"рџ’° РЎСѓРјР°: {amount} РіСЂРЅ\n"
            f"рџ’і РћРїР»Р°С‚Р°: {pay}\n"
            f"рџЏЄ {contr}\n"
        )
        if invoice:
            director_text += f"рџ”ў Р Р°С…СѓРЅРѕРє: {invoice}"

        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("вњ… РџРѕРіРѕРґРёС‚Рё", callback_data=f"director_approve_{record_id}"),
                InlineKeyboardButton("вќЊ Р’С–РґС…РёР»РёС‚Рё", callback_data=f"director_reject_{record_id}"),
            ]
        ])

        file_id   = context.user_data.get("file_id")
        file_type = context.user_data.get("file_type")

        if file_id:
            if file_type == "document":
                await context.bot.send_document(
                    chat_id=DIRECTOR_ID, document=file_id,
                    caption=director_text, parse_mode="HTML", reply_markup=keyboard
                )
            else:
                await context.bot.send_photo(
                    chat_id=DIRECTOR_ID, photo=file_id,
                    caption=director_text, parse_mode="HTML", reply_markup=keyboard
                )
        else:
            await context.bot.send_message(
                chat_id=DIRECTOR_ID, text=director_text,
                parse_mode="HTML", reply_markup=keyboard
            )

        context.user_data.clear()

    except Exception as e:
        logger.error(f"submit_for_approval error: {e}", exc_info=True)
        error_msg = str(e)[:200]
        try:
            await query.edit_message_text(
                f"вќЊ РџРѕРјРёР»РєР° Р·Р±РµСЂРµР¶РµРЅРЅСЏ:\n<code>{error_msg}</code>\n\nРЎРїСЂРѕР±СѓР№С‚Рµ С‰Рµ СЂР°Р· Р°Р±Рѕ Р·РІРµСЂРЅС–С‚СЊСЃСЏ РґРѕ Р°РґРјС–РЅР°.",
                parse_mode="HTML"
            )
        except Exception:
            await context.bot.send_message(
                chat_id=query.from_user.id,
                text=f"вќЊ РџРѕРјРёР»РєР° Р·Р±РµСЂРµР¶РµРЅРЅСЏ:\n{error_msg}\n\nРЎРїСЂРѕР±СѓР№С‚Рµ С‰Рµ СЂР°Р·."
            )

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РџРћР“РћР”Р–Р•РќРќРЇ Р”РР Р•РљРўРћР Рђ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def director_approve(query, context, record_id):
    if query.from_user.id != DIRECTOR_ID:
        await query.answer("в›” РўС–Р»СЊРєРё РґРёСЂРµРєС‚РѕСЂ РјРѕР¶Рµ РїРѕРіРѕРґР¶СѓРІР°С‚Рё.", show_alert=True)
        return
    try:
        ws = get_repairs_sheet()
        cell = ws.find(str(record_id))
        ws.update_cell(cell.row, 12, "РџРѕРіРѕРґР¶РµРЅРѕ")
        ws.update_cell(cell.row, 13, datetime.now().strftime("%d.%m.%Y %H:%M"))

        if query.message.caption:
            await query.edit_message_caption(
                query.message.caption + "\n\nвњ… <b>РџРѕРіРѕРґР¶РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>",
                parse_mode="HTML"
            )
        else:
            await query.edit_message_text(
                query.message.text + "\n\nвњ… <b>РџРѕРіРѕРґР¶РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>",
                parse_mode="HTML"
            )

        data = context.application.bot_data.get(f"repair_{record_id}", {})
        accountant_text = (
            f"рџ’і <b>Р”Рѕ РѕРїР»Р°С‚Рё вЂ” Р·Р°СЏРІРєР° #{record_id}</b>\n\n"
            f"рџ“‹ РўРёРї: {data.get('op_type','')}\n"
            f"рџљ— РњР°С€РёРЅР°: {data.get('vehicle_type','')} {data.get('vehicle','')}\n"
            f"рџ“ќ {data.get('description','')}\n"
            f"рџ’° РЎСѓРјР°: {data.get('amount','')} РіСЂРЅ\n"
            f"рџ’і РћРїР»Р°С‚Р°: {data.get('payment','')}\n"
            f"рџЏЄ {data.get('contractor','')}\n"
            f"вњ… РџРѕРіРѕРґР¶РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј"
        )
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("вњ… РћРїР»Р°С‡РµРЅРѕ", callback_data=f"accountant_paid_{record_id}")]
        ])

        file_id   = data.get("file_id")
        file_type = data.get("file_type")

        if file_id:
            if file_type == "document":
                await context.bot.send_document(
                    chat_id=ACCOUNTANT_ID, document=file_id,
                    caption=accountant_text, parse_mode="HTML", reply_markup=keyboard
                )
            else:
                await context.bot.send_photo(
                    chat_id=ACCOUNTANT_ID, photo=file_id,
                    caption=accountant_text, parse_mode="HTML", reply_markup=keyboard
                )
        else:
            await context.bot.send_message(
                chat_id=ACCOUNTANT_ID, text=accountant_text,
                parse_mode="HTML", reply_markup=keyboard
            )

        manager_id = data.get("manager_id")
        if manager_id:
            await context.bot.send_message(
                chat_id=manager_id,
                text=f"вњ… Р—Р°СЏРІРєСѓ #{record_id} <b>РїРѕРіРѕРґР¶РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>. РџРµСЂРµРґР°РЅРѕ Р±СѓС…РіР°Р»С‚РµСЂСѓ РЅР° РѕРїР»Р°С‚Сѓ.",
                parse_mode="HTML"
            )

    except Exception as e:
        logger.error(f"director_approve error: {e}")

async def director_reject(query, context, record_id):
    if query.from_user.id != DIRECTOR_ID:
        await query.answer("в›” РўС–Р»СЊРєРё РґРёСЂРµРєС‚РѕСЂ РјРѕР¶Рµ РІС–РґС…РёР»СЏС‚Рё.", show_alert=True)
        return
    try:
        ws = get_repairs_sheet()
        cell = ws.find(str(record_id))
        ws.update_cell(cell.row, 12, "Р’С–РґС…РёР»РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј")

        if query.message.caption:
            await query.edit_message_caption(
                query.message.caption + "\n\nвќЊ <b>Р’С–РґС…РёР»РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>",
                parse_mode="HTML"
            )
        else:
            await query.edit_message_text(
                query.message.text + "\n\nвќЊ <b>Р’С–РґС…РёР»РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>",
                parse_mode="HTML"
            )

        data = context.application.bot_data.get(f"repair_{record_id}", {})
        manager_id = data.get("manager_id")
        if manager_id:
            await context.bot.send_message(
                chat_id=manager_id,
                text=f"вќЊ Р—Р°СЏРІРєСѓ #{record_id} <b>РІС–РґС…РёР»РµРЅРѕ РґРёСЂРµРєС‚РѕСЂРѕРј</b>.",
                parse_mode="HTML"
            )
    except Exception as e:
        logger.error(f"director_reject error: {e}")

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РџР†Р”РўР’Р•Р Р”Р–Р•РќРќРЇ Р‘РЈРҐР“РђР›РўР•Р Рђ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def accountant_paid(query, context, record_id):
    if query.from_user.id != ACCOUNTANT_ID:
        await query.answer("в›” РўС–Р»СЊРєРё Р±СѓС…РіР°Р»С‚РµСЂ РјРѕР¶Рµ РїС–РґС‚РІРµСЂРґР¶СѓРІР°С‚Рё РѕРїР»Р°С‚Сѓ.", show_alert=True)
        return
    try:
        ws = get_repairs_sheet()
        cell = ws.find(str(record_id))
        ws.update_cell(cell.row, 12, "РћРїР»Р°С‡РµРЅРѕ")
        ws.update_cell(cell.row, 14, datetime.now().strftime("%d.%m.%Y %H:%M"))

        if query.message.caption:
            await query.edit_message_caption(
                query.message.caption + "\n\nвњ… <b>РћРїР»Р°С‡РµРЅРѕ</b>",
                parse_mode="HTML"
            )
        else:
            await query.edit_message_text(
                query.message.text + "\n\nвњ… <b>РћРїР»Р°С‡РµРЅРѕ</b>",
                parse_mode="HTML"
            )

        data = context.application.bot_data.get(f"repair_{record_id}", {})
        manager_id = data.get("manager_id")
        if manager_id:
            await context.bot.send_message(
                chat_id=manager_id,
                text=f"вњ… Р—Р°СЏРІРєСѓ #{record_id} <b>РѕРїР»Р°С‡РµРЅРѕ</b> Р±СѓС…РіР°Р»С‚РµСЂРѕРј.",
                parse_mode="HTML"
            )

        op = data.get("op_type", "")
        if op == "purchase":
            name  = data.get("stock_name", "")
            qty   = float(data.get("stock_qty", 0) or 0)
            unit  = data.get("stock_unit", "")
            amount = float(data.get("amount", 0) or 0)
            price_per_unit = round(amount / qty, 2) if qty else 0
            if name and qty:
                update_stock(name, unit, qty, price_per_unit)

    except Exception as e:
        logger.error(f"accountant_paid error: {e}")

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# РЎРџРРЎРђРќРќРЇ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def submit_writeoff(query, context):
    try:
        item    = context.user_data.get("stock_item", {})
        qty     = context.user_data.get("writeoff_qty", 0)
        vehicle = context.user_data.get("vehicle", "")
        vtype   = context.user_data.get("vehicle_type", "")

        name  = item.get("РџРѕР·РёС†С–СЏ", "")
        unit  = item.get("РћРґРёРЅРёС†СЏ", "")
        price = float(item.get("Р¦С–РЅР° Р·Р° РѕРґРёРЅРёС†СЋ", 0) or 0)
        total = round(qty * price, 2)

        context.user_data["step"] = "confirm_writeoff"
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("вњ… РўР°Рє, РІС–РґРїСЂР°РІРёС‚Рё", callback_data="confirm_data"),
                InlineKeyboardButton("вњЏпёЏ Р’РёРїСЂР°РІРёС‚Рё", callback_data="back_main"),
            ]
        ])
        await query.edit_message_text(
            f"<b>РџРµСЂРµРІС–СЂС‚Рµ РґР°РЅС– СЃРїРёСЃР°РЅРЅСЏ:</b>\n\n"
            f"рџ“¤ РўРёРї: РЎРїРёСЃР°РЅРЅСЏ\n"
            f"рџ“¦ РџРѕР·РёС†С–СЏ: {name}\n"
            f"рџ“Љ РљС–Р»СЊРєС–СЃС‚СЊ: {qty} {unit}\n"
            f"рџљ— РњР°С€РёРЅР°: {vtype} {vehicle}\n"
            f"рџ’° Р’Р°СЂС‚С–СЃС‚СЊ: {total} РіСЂРЅ\n\n"
            f"Р’С–РґРїСЂР°РІРёС‚Рё РґРёСЂРµРєС‚РѕСЂСѓ РЅР° РїРѕРіРѕРґР¶РµРЅРЅСЏ?",
            parse_mode="HTML",
            reply_markup=keyboard
        )
        context.user_data.update({
            "op_type": "writeoff",
            "description": f"РЎРїРёСЃР°РЅРЅСЏ Р·С– СЃРєР»Р°РґСѓ: {name} {qty} {unit}",
            "amount": str(total),
            "payment": "вЂ”",
            "contractor": "РЎРєР»Р°Рґ",
            "invoice": "",
            "date": datetime.now().strftime("%d.%m.%Y"),
        })

    except Exception as e:
        logger.error(f"submit_writeoff error: {e}")

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р—Р’Р†Рў /report
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
async def report_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in MANAGER_IDS + [DIRECTOR_ID]:
        return
    try:
        ws = get_repairs_sheet()
        records = ws.get_all_records()
        current_month = datetime.now().strftime("%m.%Y")

        paid = [r for r in records
                if r.get("РЎС‚Р°С‚СѓСЃ") == "РћРїР»Р°С‡РµРЅРѕ"
                and current_month in str(r.get("Р”Р°С‚Р° РїРѕРґР°С‡С–", ""))]

        if not paid:
            await update.message.reply_text(f"Р—Р° {current_month} РѕРїР»Р°С‡РµРЅРёС… Р·Р°СЏРІРѕРє РЅРµРјР°С”.")
            return

        by_vehicle = {}
        for r in paid:
            v = r.get("РњР°С€РёРЅР°", "РЅРµРІС–РґРѕРјРѕ")
            amount = float(str(r.get("РЎСѓРјР°", 0)).replace(" ", "") or 0)
            by_vehicle[v] = by_vehicle.get(v, 0) + amount

        total = sum(by_vehicle.values())
        top5  = sorted(by_vehicle.items(), key=lambda x: x[1], reverse=True)[:5]
        cash  = sum(float(str(r.get("РЎСѓРјР°",0)).replace(" ","") or 0)
                    for r in paid if r.get("Р¤РѕСЂРјР° РѕРїР»Р°С‚Рё") == "РіРѕС‚С–РІРєР°")
        bank  = total - cash

        lines = [f"рџ“Љ <b>Р—РІС–С‚ Р·Р° {current_month}</b>\n"]
        lines.append(f"рџ’° Р—Р°РіР°Р»СЊРЅР° СЃСѓРјР°: <b>{total:,.0f} РіСЂРЅ</b>")
        lines.append(f"рџ’µ Р“РѕС‚С–РІРєР°: {cash:,.0f} РіСЂРЅ")
        lines.append(f"рџЏ¦ Р‘РµР·РЅР°Р»: {bank:,.0f} РіСЂРЅ")
        lines.append(f"рџ“‹ Р—Р°СЏРІРѕРє: {len(paid)}\n")
        lines.append("<b>РўРћРџ-5 РјР°С€РёРЅ:</b>")
        for i, (v, amt) in enumerate(top5, 1):
            lines.append(f"{i}. {v} вЂ” {amt:,.0f} РіСЂРЅ")

        await update.message.reply_text("\n".join(lines), parse_mode="HTML")

    except Exception as e:
        logger.error(f"report_cmd error: {e}")
        await update.message.reply_text("РџРѕРјРёР»РєР° РїСЂРё С„РѕСЂРјСѓРІР°РЅРЅС– Р·РІС–С‚Сѓ.")

# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
# Р—РђРџРЈРЎРљ
# в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђ
def main():
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN РЅРµ РІСЃС‚Р°РЅРѕРІР»РµРЅРѕ! Р”РѕРґР°Р№С‚Рµ Р·РјС–РЅРЅСѓ BOT_TOKEN РІ Railway.")
        return

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("new",   new_cmd))
    app.add_handler(CommandHandler("report", report_cmd))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(
        filters.Document.ALL | filters.PHOTO,
        handle_file
    ))
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        handle_text
    ))

    logger.info("Repair bot started...")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
